import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_workbook(filename):


    wb=xl.load_workbook(filename)

    sheet=wb['Sheet1']

    # cell =sheet['a1']

    # cell =sheet.cell(1,1)
    # print(cell.value)


    # print(f"the length of the row is {sheet.max_row}")

    sheet.cell(1,4).value="10% discount"



    for row in range(2,sheet.max_row+1):
        get_cell=sheet.cell(row,3)
        # print(get_cell.value)
        discount_price=get_cell.value*0.9
        discount_price_cell=sheet.cell(row,4)
        discount_price_cell.value=discount_price


    values=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)

    chart=BarChart()

    chart.add_data(values)

    sheet.add_chart(chart,'e2')

    wb.save(filename)


process_workbook('transactions.xlsx')



